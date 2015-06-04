from openpyxl import load_workbook


longitudRecuadro = 383
wb = load_workbook(filename = 'work.xlsx')

Secuencia = wb.get_sheet_by_name('Secuencia')
DESIGNER = wb.get_sheet_by_name('DESIGNER')

Hilo = ""
conector1 = ""
via1 = ""
conector2 = ""
via2 = ""
color = ""


print "hola"


for currentRow in range(longitudRecuadro):

	#if DESIGNER.cell(row=currentRow,column=2).value != "":
	Hilo = str(DESIGNER.cell(row=currentRow,column=1).value)+"-"+str(DESIGNER.cell(row=currentRow,column=2).value)
	conector1 = str(DESIGNER.cell(row=currentRow,column=5).value)
	via1 = str(DESIGNER.cell(row=currentRow,column=6).value)
	conector2 = str(DESIGNER.cell(row=currentRow,column=7).value)
	via2 = str(DESIGNER.cell(row=currentRow,column=8).value)
	if DESIGNER.cell(row=currentRow,column=4).value != "":
		color = str(DESIGNER.cell(row=currentRow,column=3).value)+"-"+str(DESIGNER.cell(row=currentRow,column=4).value)
	else:
		color = DESIGNER.cell(row=currentRow,column=3).value
	print str(currentRow) + " - "+ color

	Hilo = ""
	conector1 = ""
	via1 = ""
	conector2 = ""
	via2 = ""
	color = ""

'''
	conector1 = DESIGNER['A2'].value
	via1 = DESIGNER['A2'].value
	conector2 = DESIGNER['A2'].value
	via2 = DESIGNER['A2'].value
	'''